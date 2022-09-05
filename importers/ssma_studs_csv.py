import csv
import openpyxl
import re

from pathlib import Path

# connect to csv file
with open('./output/stud_sections.csv', 'w', newline='') as csvfile:
    headers = ('description', 'design_thickness', 'Fy', 'area', 'weight',
                    'Ix', 'Sx', 'Rx', 'Iy', 'Ry', 'Ixe', 'Sxe', 'Mal', 'Mad',
                    'Vag', 'Vanet', 'J', 'Cw', 'Xo', 'm', 'Ro', 'Beta', 'Lu',
                    'notes')
    writer = csv.DictWriter(csvfile, fieldnames=headers)
    writer.writeheader()
    
    # footnotes from SSMA stud section properties tables
    footnote_1 = 'FOOTNOTE 1 Web height-to-thickness ratio exceeds 200. Web stiffeners are required at all support points and concentrated loads.'
    footnote_2 = 'FOOTNOTE 2 Allowable moment includes cold work of forming.'
    footnote_3 = 'FOOTNOTE 3 Where web height-to-thickness ratio exceeds 260 or flange width-to-thickness ratio exceeds 60, effective properties are not calculated. See AISI S100 Section B1. Application of these products in a non-composite design shall be approved by a design professional.'

    # open excel file
    xlsx_file = Path('sources', 'SSMA_Product_Technical_Guide_Studs.xlsx')
    wb = openpyxl.load_workbook(xlsx_file)
    sheet = wb['Tables Combined']
  
    # sheet.iter_rows(12,504,1,23) # all stud data
    for row in sheet.iter_rows(12, 504, 1, 23, True):
        print(row)
        colIndex = 0  # so we can keep track of which header we're on
        rowDict = {}
        notes = None
        notes_list = []  # start with no notes
        for col in row:
            val = col
            if type(col) == str:
                # see if it's blank or '-'
                if col == '-':
                    val = None
                else:
                    # see if there are footnotes
                    # for example, '550S125-18 1, 3'
                    splits = re.split(r'[\ ,]',
                                      col)  # ['550S125-18', '1', '', '3']
                    filtered_splits = list(filter(
                        None, splits))  # ['550S125-18', '1', '3']
                    val = filtered_splits[0]
                    footnotes = filtered_splits[1:]  # ['1','3']
                    for footnote in footnotes:
                        # for example: 'description FOOTNOTE 1 Web height-to-thickness...'
                        if footnote == '1':
                            notes_list.append(headers[colIndex] + ' ' +
                                              footnote_1)
                        if footnote == '2':
                            notes_list.append(headers[colIndex] + ' ' +
                                              footnote_2)
                        if footnote == '3':
                            notes_list.append(headers[colIndex] + ' ' +
                                              footnote_3)
            rowDict[(headers[colIndex])] = val
            colIndex = colIndex + 1  # next column

        # compile footers, if any
        if notes_list.__len__() > 0:
            rowDict['notes'] = "; ".join(notes_list)
    
        writer.writerow(rowDict)

